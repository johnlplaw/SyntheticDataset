# This script is for extracting the data from the Excel file and insert it into database.
# Only the selected lables are selected.
# The source data are from the excel file.
# 2024-01-07 - Initial Coding

import openpyxl
import commons.mysql.mysqlHelper as sqlHelper

class SrcData:
    oriTxt = ""
    labelTxt = ""
    cleanedTxt = ""

    def __init__(self, oriTxt, labelTxt, cleanedTxt):
        self.oriTxt = oriTxt
        self.labelTxt = labelTxt
        self.cleanedTxt = cleanedTxt

def formatLabelTxt(input):
    if input :
        # trim the text
        input = input.strip()
        # lower case
        input = input.lower()
    return input

def handleSheet1(sheets):
    # for sheet 1 - emotion-in-text
    # col1 = text
    # col2 = emotion
    # selected label = happy, anger, sadness, surprise, fear

    acceptedLabel = ['happy', 'anger', 'sadness', 'surprise', 'fear']
    ws = wb[sheets[0]]
    column1 = ws['A']
    column2 = ws['B']
    datalist = []
    for row in range(1, ws.max_row):
        oriTxt = column1[row].value
        labelTxt = column2[row].value

        labelTxt = formatLabelTxt(labelTxt)
        if labelTxt in acceptedLabel:
            obj = SrcData(oriTxt, labelTxt, None)
            datalist.append(obj)

    print("Selected data 1 size:")
    print(len(datalist))
    return datalist


def handleSheet2(sheets):
    # for sheet 2 - Emotion Classification NLP
    # col1 = text
    # col2 = emotion
    # selected label = happy, anger, sadness, surprise, fear

    acceptedLabel = ['joy', 'anger', 'sadness', 'surprise', 'fear']
    ws = wb[sheets[1]]
    column1 = ws['A']
    column2 = ws['B']
    datalist = []
    for row in range(1, ws.max_row):
        oriTxt = column1[row].value
        labelTxt = column2[row].value

        labelTxt = formatLabelTxt(labelTxt)
        if labelTxt in acceptedLabel:
            obj = SrcData(oriTxt, labelTxt, None)
            datalist.append(obj)

    print("Selected data 2 size:")
    print(len(datalist))
    return datalist

def handleSheet3(sheets):
    # for sheet 3 - Emotion Detection from Text
    # col3 = text
    # col2 = emotion
    # selected label = happiness, anger, sadness, surprise, worry

    acceptedLabel = ['happiness', 'anger', 'sadness', 'surprise', 'worry', 'neutral']

    ws = wb[sheets[2]]
    datalist = []

    column1 = ws['A']
    column2 = ws['B']
    column3 = ws['C']

    i = 0
    for row in range(1, ws.max_row):
        labelTxt = column2[row].value
        oriTxt = column3[row].value

        labelTxt = formatLabelTxt(labelTxt)
        if labelTxt in acceptedLabel:
            obj = SrcData(oriTxt, labelTxt, None)
            datalist.append(obj)

    print("Selected data 3 size:")
    print(len(datalist))
    return datalist


def handleSheet4(sheets):
    # for sheet 4 - cleaned Emotion Extraction data
    # col3 = clean text
    # col2 = text
    # col1 = emotion
    # selected label = happy, angry, disappointed, surprise, worry

    acceptedLabel = ['happy', 'angry', 'disappointed', 'surprise', 'worry']
    ws = wb[sheets[3]]
    datalist = []

    column1 = ws['A']
    column2 = ws['B']
    column3 = ws['C']

    i = 0
    for row in range(1, ws.max_row):
        labelTxt = column1[row].value
        oriTxt = column3[row].value
        cleanedTxt = column2[row].value

        labelTxt = formatLabelTxt(labelTxt)
        if labelTxt in acceptedLabel:
            obj = SrcData(oriTxt, labelTxt, cleanedTxt)
            datalist.append(obj)

    print("Selected data 4 size:")
    print(len(datalist))
    return datalist


# Define variable to load the wb
wb = openpyxl.load_workbook("../data/SourceDataSets.xlsx")
sheets = wb.sheetnames

list1 = handleSheet1(sheets)
list2 = handleSheet2(sheets)
list3 = handleSheet3(sheets)
list4 = handleSheet4(sheets)

selectedList = list1 + list2 + list3 + list4

print(len(selectedList))

print("Insert into mysql - start")

conn = sqlHelper.get_mysql_conn()
mycursor = conn.cursor()
sql = "INSERT INTO Synth_text (label, oritxt, cleanedtxt) VALUES (%s, %s, %s)"

for txtdata in selectedList:
    val = ( txtdata.labelTxt, txtdata.oriTxt, txtdata.cleanedTxt)
    mycursor.execute(sql, val)

conn.commit()
print(mycursor.rowcount, "record inserted.")

print("Insert into mysql - end")










